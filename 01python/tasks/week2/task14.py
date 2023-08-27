#! /usr/bin/python3

# Python program to parse header file and read
# all prototypes of function and insert it into
# excel sheet with unique ID start with IDX0

import openpyxl
import re

wb = openpyxl.Workbook()
ws = wb.create_sheet("Functions")
#ws = wb.active()
ws['A1'] = 'Function'
ws['B1'] = 'ID'
count = 0
with open('DIO_INTERFACE.h','+r') as file_h:
    for line in file_h.readlines():
        if re.search("[s-v].*\(.*\)" ,line):
            id_value = "IDX0"+str(count)
            ws.append([line,id_value])
            count = count+1
wb.save("Functions.xlsx")